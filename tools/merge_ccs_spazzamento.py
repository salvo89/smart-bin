# -*- coding: utf-8 -*-
"""Merge CCS Chieri/Carmagnola spazzamento (bin 6) into existing zone calendars.

Sources (XLSX on ccs.to.it/calendari-spazzamento):
- Chieri: 2026_CHIERI_CALENDARIO_SPAZZAMENTI.xlsx
- Carmagnola: CARMAGNOLA_MECCANIZZATO.xlsx / CARMAGNOLA_MANUALE.xlsx

Does not create new calendars: merges Spazzamento into existing .h zone files
by matching Escilo vie names to the street→zona map in the workbook.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

try:
    import openpyxl
except ImportError:  # pragma: no cover
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

from tools.covar14_pdf_to_h import write_year_file  # noqa: E402
from tools.download_safe import download_if_needed  # noqa: E402

BIN_SPA = 6
ENTRY_RE = re.compile(
    r"\{(\d{4})\s*,\s*(\d{1,2})\s*,\s*(\d{1,2})\s*,\s*(-?\d+)\s*\}"
)

CHIERI_URL = (
    "https://www.ccs.to.it/flex/Extensions/appCCSCalendario/pages/"
    "serveDownload.php?a=2026&f=2026_CHIERI_CALENDARIO_SPAZZAMENTI.xlsx&t=spazzamento"
)
CARM_MECC_URL = (
    "https://www.ccs.to.it/flex/Extensions/appCCSCalendario/pages/"
    "serveDownload.php?a=2026&f=CARMAGNOLA_MECCANIZZATO.xlsx&t=spazzamento"
)
CARM_MAN_URL = (
    "https://www.ccs.to.it/flex/Extensions/appCCSCalendario/pages/"
    "serveDownload.php?a=2026&f=CARMAGNOLA_MANUALE.xlsx&t=spazzamento"
)


def norm_street(s: str) -> str:
    s = str(s).upper()
    for a, b in (
        ("À", "A"),
        ("È", "E"),
        ("É", "E"),
        ("Ì", "I"),
        ("Ò", "O"),
        ("Ù", "U"),
        ("�", "A"),
        ("'", ""),
        ("’", ""),
    ):
        s = s.replace(a, b)
    s = re.sub(r"^VIA\s+|^VIALE\s+|^CORSO\s+|^PIAZZA\s+|^STRADA\s+|^LARGO\s+", "", s)
    # drop trailing person-name reorder noise: keep tokens
    return re.sub(r"\s+", " ", s).strip()


def load_entries(path: Path) -> list[tuple[int, int, int, int]]:
    text = path.read_text(encoding="utf-8")
    return [(int(y), int(m), int(d), int(b)) for y, m, d, b in ENTRY_RE.findall(text)]


def merge_spa(
    existing: list[tuple[int, int, int, int]],
    spa: list[tuple[int, int, int, int]],
) -> list[tuple[int, int, int, int]]:
    kept = [e for e in existing if e[3] != BIN_SPA]
    return sorted(set(kept) | set(spa))


def parse_calendar_sheet(ws) -> dict[str, list[tuple[int, int, int, int]]]:
    """zona_code -> list of (y,m,d,6) where cell is 'x'."""
    dates: list[tuple[int, int, int, int]] = []  # col, y, m, d
    for c in range(2, (ws.max_column or 0) + 1):
        v = ws.cell(1, c).value
        if isinstance(v, datetime):
            dates.append((c, v.year, v.month, v.day))
        elif isinstance(v, str) and re.match(r"\d{4}-\d{2}-\d{2}", v):
            dt = datetime.fromisoformat(v[:10])
            dates.append((c, dt.year, dt.month, dt.day))
    out: dict[str, list[tuple[int, int, int, int]]] = {}
    for r in range(2, (ws.max_row or 0) + 1):
        z = ws.cell(r, 1).value
        if z is None:
            continue
        # skip weekday header row
        if str(z).strip().upper() in {"ZONA", "ZONE"}:
            continue
        code = str(z).strip()
        entries: list[tuple[int, int, int, int]] = []
        for c, y, m, d in dates:
            cell = ws.cell(r, c).value
            if cell is None:
                continue
            if str(cell).strip().lower() in {"x", "1", "sì", "si", "true"}:
                entries.append((y, m, d, BIN_SPA))
        if entries:
            out[code] = entries
    return out


def parse_street_zones(ws, zona_col: int = 3) -> dict[str, set[str]]:
    street_zones: dict[str, set[str]] = defaultdict(set)
    # detect header
    start = 1
    for r in range(1, min(5, (ws.max_row or 1) + 1)):
        v = str(ws.cell(r, 1).value or "").upper()
        if "NOME" in v or "VIA" in v:
            start = r + 1
            # find zona col
            for c in range(1, (ws.max_column or 0) + 1):
                h = str(ws.cell(r, c).value or "").upper()
                if "ZONA" in h:
                    zona_col = c
            break
    for r in range(start, (ws.max_row or 0) + 1):
        name = ws.cell(r, 1).value
        zona = ws.cell(r, zona_col).value
        if not name or not zona:
            continue
        for code in re.split(r"[,\s;/]+", str(zona).replace("-", " ")):
            # codes like D14 or D02 — keep tokens with letter+digits
            code = code.strip().upper()
            if re.fullmatch(r"[A-Z]\d{1,3}[A-Z]?", code) or re.fullmatch(r"[A-Z]\d+", code):
                street_zones[norm_street(name)].add(code)
            elif re.fullmatch(r"[A-Z]\d{2}", code):
                street_zones[norm_street(name)].add(code)
        # also handle "D02-D03-D21" without splitting wrongly — already split by -
        raw = str(zona).upper()
        for code in re.findall(r"[A-Z]\d{1,3}", raw):
            street_zones[norm_street(name)].add(code)
    return street_zones


def lookup_codes(street: str, street_zones: dict[str, set[str]]) -> set[str]:
    n = norm_street(street)
    if n in street_zones:
        return set(street_zones[n])
    # token-ish fuzzy: exact containment
    hits: set[str] = set()
    for k, v in street_zones.items():
        if n == k or (len(n) >= 6 and (n in k or k in n)):
            hits |= v
    return hits


def escilo_calendars(index_path: Path, comune_id: str) -> dict[str, list[str]]:
    data = json.loads(index_path.read_text(encoding="utf-8"))
    comune = next(c for c in data["comuni"] if c["id"] == comune_id)
    by_cal: dict[str, list[str]] = defaultdict(list)
    for v in comune["vie"]:
        by_cal[v["calendar"]].append(v["name"])
    return by_cal


def dates_for_codes(
    codes: set[str], calendars: list[dict[str, list[tuple[int, int, int, int]]]]
) -> list[tuple[int, int, int, int]]:
    out: set[tuple[int, int, int, int]] = set()
    for cal in calendars:
        for code in codes:
            out.update(cal.get(code, []))
    return sorted(out)


def merge_comune(
    *,
    comune_id: str,
    comune_name: str,
    year: int,
    outdir: Path,
    by_cal: dict[str, list[str]],
    street_zones: dict[str, set[str]],
    cal_maps: list[dict[str, list[tuple[int, int, int, int]]]],
    dry_run: bool,
    min_matched: int = 3,
) -> tuple[int, int]:
    merged = 0
    skipped = 0
    for cal_path, streets in sorted(by_cal.items()):
        slug = cal_path.replace("calendars/", "")
        path = outdir / f"{slug}-{year}.h"
        if not path.exists():
            print(f"skip missing {path.name}")
            skipped += 1
            continue
        code_votes: Counter[str] = Counter()
        matched = 0
        for st in streets:
            codes = lookup_codes(st, street_zones)
            if codes:
                matched += 1
                for c in codes:
                    code_votes[c] += 1
        if matched < min_matched or not code_votes:
            print(f"skip {slug}: matched streets {matched}/{len(streets)}")
            skipped += 1
            continue
        # One dominant spazzamento code per Escilo zone (avoid near-daily unions).
        top = code_votes.most_common()
        selected = {top[0][0]}
        if len(top) > 1 and top[1][1] >= max(2, int(0.8 * top[0][1])):
            selected.add(top[1][0])
        spa = dates_for_codes(selected, cal_maps)
        spa = [e for e in spa if e[0] == year]
        if len(spa) < 5:
            print(f"skip {slug}: too few spa days {len(spa)} codes={sorted(selected)}")
            skipped += 1
            continue
        existing = load_entries(path)
        before = sum(1 for e in existing if e[3] == BIN_SPA)
        new_entries = merge_spa(existing, spa)
        after = sum(1 for e in new_entries if e[3] == BIN_SPA)
        print(
            f"merge {path.name}: matched={matched}/{len(streets)} "
            f"codes={sorted(selected)} votes={dict(top[:3])} spa {before}->{after}"
        )
        if not dry_run:
            first = path.read_text(encoding="utf-8").splitlines()[0]
            zone_label = slug
            m = re.search(r"\(([^)]+)\)", first)
            # Prefer "Zona …" from header if present
            zm = re.search(r"^(?://\s*)?[^(]+?\s+(Zona\s+[^(]+)", first, re.I)
            if zm:
                zone_label = zm.group(1).strip()
            elif "Zona" in first:
                zone_label = first.split("(")[0].replace("//", "").strip()
                # drop comune name prefix
                zone_label = re.sub(rf"^{re.escape(comune_name)}\s+", "", zone_label)
            write_year_file(
                out_path=path,
                comune_name=comune_name,
                zone_label=zone_label or slug,
                provider="CCS",
                addresses=[],
                year=year,
                entries=new_entries,
            )
        merged += 1
    return merged, skipped


def process_chieri(outdir: Path, work: Path, year: int, dry_run: bool) -> None:
    xlsx = download_if_needed(CHIERI_URL, work / "chieri.xlsx")
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    # Pair elenco+calendario per servizio, with disjoint code namespaces.
    man_streets: dict[str, set[str]] = defaultdict(set)
    mec_streets: dict[str, set[str]] = defaultdict(set)
    man_cal: dict[str, list[tuple[int, int, int, int]]] = {}
    mec_cal: dict[str, list[tuple[int, int, int, int]]] = {}

    for name in wb.sheetnames:
        ws = wb[name]
        up = name.upper()
        if "ELENCO" in up and "MANUALE" in up:
            for k, v in parse_street_zones(ws).items():
                man_streets[k] |= {f"man:{c}" for c in v}
        elif "ELENCO" in up and "MECCA" in up:
            for k, v in parse_street_zones(ws).items():
                mec_streets[k] |= {f"mec:{c}" for c in v}
        elif "CALENDARIO" in up and "MANUALE" in up:
            raw = parse_calendar_sheet(ws)
            man_cal = {f"man:{k}": v for k, v in raw.items()}
        elif "CALENDARIO" in up and "MECCA" in up:
            raw = parse_calendar_sheet(ws)
            mec_cal = {f"mec:{k}": v for k, v in raw.items()}

    # Prefer meccanizzato street map; fill gaps from manuale.
    street_zones: dict[str, set[str]] = defaultdict(set)
    for k, v in mec_streets.items():
        street_zones[k] |= v
    for k, v in man_streets.items():
        if k not in street_zones:
            street_zones[k] |= v

    cal_maps = [mec_cal, man_cal]
    print(
        f"Chieri: streets={len(street_zones)} "
        f"mec_codes={len(mec_cal)} man_codes={len(man_cal)}"
    )
    by_cal = escilo_calendars(outdir / "index.json", "chieri")
    merged, skipped = merge_comune(
        comune_id="chieri",
        comune_name="Chieri",
        year=year,
        outdir=outdir,
        by_cal=by_cal,
        street_zones=street_zones,
        cal_maps=cal_maps,
        dry_run=dry_run,
    )
    print(f"Chieri done: merged={merged} skipped={skipped}")


def process_carmagnola(outdir: Path, work: Path, year: int, dry_run: bool) -> None:
    street_zones: dict[str, set[str]] = defaultdict(set)
    cal_maps: list[dict[str, list[tuple[int, int, int, int]]]] = []
    for label, url in (
        ("carm_mecc", CARM_MECC_URL),
        ("carm_man", CARM_MAN_URL),
    ):
        try:
            xlsx = download_if_needed(url, work / f"{label}.xlsx", force=False)
        except Exception as exc:  # noqa: BLE001
            print(f"Carmagnola skip {label}: {exc}")
            continue
        wb = openpyxl.load_workbook(xlsx, data_only=True)
        print(f"Carmagnola {label} sheets={wb.sheetnames}")
        for name in wb.sheetnames:
            ws = wb[name]
            up = name.upper()
            # Heuristic: calendars are wide; elenchi are tall with few cols
            if (ws.max_column or 0) >= 30:
                cal_maps.append(parse_calendar_sheet(ws))
            if (ws.max_column or 0) <= 10 and (ws.max_row or 0) >= 20:
                for k, v in parse_street_zones(ws).items():
                    street_zones[k] |= v
            # also try both always
            if "ELENCO" in up or "VIE" in up:
                for k, v in parse_street_zones(ws).items():
                    street_zones[k] |= v
            if "CALENDARIO" in up:
                cal_maps.append(parse_calendar_sheet(ws))
    if not cal_maps:
        print("Carmagnola: no calendar sheets parsed")
        return
    print(
        f"Carmagnola: streets={len(street_zones)} cal_sheets={len(cal_maps)} "
        f"codes={sum(len(m) for m in cal_maps)}"
    )
    by_cal = escilo_calendars(outdir / "index.json", "carmagnola")
    merged, skipped = merge_comune(
        comune_id="carmagnola",
        comune_name="Carmagnola",
        year=year,
        outdir=outdir,
        by_cal=by_cal,
        street_zones=street_zones,
        cal_maps=cal_maps,
        dry_run=dry_run,
        min_matched=2,
    )
    print(f"Carmagnola done: merged={merged} skipped={skipped}")


def update_sources(sources_path: Path, dry_run: bool) -> None:
    data = json.loads(sources_path.read_text(encoding="utf-8"))
    for cid, note, url in (
        (
            "chieri",
            "Spazzamento 2026 (CCS XLSX) mergiato come bin Spazzamento nelle zone esistenti.",
            CHIERI_URL,
        ),
        (
            "carmagnola",
            "Spazzamento 2026 (CCS XLSX meccanizzato/manuale) mergiato come bin Spazzamento.",
            CARM_MECC_URL,
        ),
    ):
        comune = next(c for c in data["comuni"] if c["id"] == cid)
        notes = list(comune.get("notes") or [])
        if note not in notes:
            notes.append(note)
        pdfs = list(comune.get("pdfs") or [])
        if not any(p.get("url") == url for p in pdfs):
            pdfs.append({"year": 2026, "label": "Spazzamento 2026", "url": url})
        if dry_run:
            continue
        comune["notes"] = notes
        comune["pdfs"] = pdfs
    if dry_run:
        print("[dry-run] sources notes/pdfs would update")
        return
    sources_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print("sources.json updated for Chieri/Carmagnola spazzamento")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--year", type=int, default=2026)
    ap.add_argument("--outdir", type=Path, default=ROOT / "docs" / "calendars")
    ap.add_argument("--work", type=Path, default=ROOT / "tmp_calendars" / "ccs_spazz")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--only", choices=["chieri", "carmagnola", "all"], default="all")
    args = ap.parse_args()
    args.work.mkdir(parents=True, exist_ok=True)

    if args.only in {"chieri", "all"}:
        process_chieri(args.outdir, args.work, args.year, args.dry_run)
    if args.only in {"carmagnola", "all"}:
        process_carmagnola(args.outdir, args.work, args.year, args.dry_run)
    update_sources(args.outdir / "sources.json", args.dry_run)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
